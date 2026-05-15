"""XLSX via openpyxl: sheets → markdown tables (native Python, no kordoc)."""

from __future__ import annotations

import logging
from importlib.metadata import PackageNotFoundError, version
from io import BytesIO

from openpyxl import load_workbook

from app.adapters.parser_protocol import ParseRequest, ParseResult, ParserClient

log = logging.getLogger("contexthub.parser.xlsx")


def _openpyxl_version() -> str:
    try:
        return version("openpyxl")
    except PackageNotFoundError:
        return "unknown"


class XlsxOpenpyxlParser(ParserClient):
    def parse(self, request: ParseRequest) -> ParseResult:
        try:
            wb = load_workbook(BytesIO(request.file_bytes), read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError(f"Invalid or unreadable XLSX: {exc}") from exc

        lines: list[str] = []
        blocks: list[dict[str, object]] = []
        sheet_count = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_count += 1
            lines.append(f"## {sheet_name}")
            blocks.append({"type": "sheet", "name": sheet_name})

            rows_iter = ws.iter_rows(values_only=True)
            first_row = next(rows_iter, None)
            if first_row is None:
                lines.append("_(empty sheet)_")
                continue

            header = [str(c).strip() if c is not None else "" for c in first_row]
            header = [h if h else f"col{i + 1}" for i, h in enumerate(header)]
            lines.append("| " + " | ".join(header) + " |")
            lines.append("| " + " | ".join("---" for _ in header) + " |")

            row_count = 0
            for row in rows_iter:
                cells = [str(c).strip() if c is not None else "" for c in row]
                if not any(cells):
                    continue
                while len(cells) < len(header):
                    cells.append("")
                lines.append("| " + " | ".join(cells[: len(header)]) + " |")
                row_count += 1
                if row_count <= 3:
                    blocks.append(
                        {
                            "type": "row",
                            "sheet": sheet_name,
                            "preview": cells[: min(4, len(cells))],
                        }
                    )

        wb.close()
        md = "\n\n".join(lines)
        ver = _openpyxl_version()
        meta: dict[str, object] = {
            "engine": "openpyxl",
            "openpyxl_version": ver,
            "original_filename": request.original_filename,
            "sheet_count": sheet_count,
        }
        log.info(
            "openpyxl parsed filename=%r sheets=%s markdown_chars=%s",
            request.original_filename,
            sheet_count,
            len(md),
        )
        return ParseResult(
            markdown_text=md,
            blocks_json=blocks,
            metadata_json=meta,
            page_count=sheet_count if sheet_count else None,
            parser_version=f"openpyxl-{ver}",
            parser_name="openpyxl",
        )
